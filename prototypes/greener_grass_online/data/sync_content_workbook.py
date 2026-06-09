from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path("/Users/lydukhanhhan/Documents/MA_ResearchLog")
DATA_DIR = PROJECT_ROOT / "prototypes" / "greener_grass" / "data"
DEFAULT_WORKBOOK = PROJECT_ROOT / "outputs" / "greener_grass_content_templates" / "greener_grass_content_templates.xlsx"

SHEET_TO_CSV = {
    "space_data_template": DATA_DIR / "space_data_template.csv",
    "life_cards_template": DATA_DIR / "life_cards_template.csv",
    "era_cards_template": DATA_DIR / "era_cards_template.csv",
    "occupations_template": DATA_DIR / "occupations_template.csv",
    "hobbies_template": DATA_DIR / "hobbies_template.csv",
}


def cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def trimmed_matrix(worksheet) -> list[list[str]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [cell_to_text(cell) for cell in rows[0]]
    while header and header[-1] == "":
        header.pop()
    width = len(header)
    if width == 0:
        return []

    matrix = [header]
    for row in rows[1:]:
        values = [cell_to_text(cell) for cell in row[:width]]
        if any(value != "" for value in values):
            matrix.append(values)

    return matrix


def export_sheet(worksheet, target_csv: Path) -> int:
    matrix = trimmed_matrix(worksheet)
    if not matrix:
        raise ValueError(f"Sheet '{worksheet.title}' is empty or missing headers.")

    with target_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(matrix)

    return max(0, len(matrix) - 1)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync the GREENER GRASS content workbook back into CSV templates."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help="Path to the .xlsx workbook to export from.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=False)

    for sheet_name, target_csv in SHEET_TO_CSV.items():
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Workbook is missing expected sheet: {sheet_name}")
        count = export_sheet(workbook[sheet_name], target_csv)
        print(f"{sheet_name} -> {target_csv} ({count} data rows)")


if __name__ == "__main__":
    main()
